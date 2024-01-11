import os

from selenium.common import TimeoutException
from selenium.webdriver.common.by import By
from src.utils.WaitingUtils import WaitingUtils
from src.utils import Utils
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from selenium import webdriver
import datetime


class Login:

    def login(self):

        wait = WaitingUtils
        post = input("Introdu link-ul postarii: ")
        pause = input("Introdu numarul de minute intre postari: ")
        pause_min = int(pause)
        current_file_path = os.path.abspath(__file__)
        parent_directory = os.path.dirname(current_file_path)
        project_root = os.path.dirname(parent_directory)
        file_path = os.path.join(project_root, "resources", "users.xlsx")
        run_path = os.path.join(project_root, "reports")
        report_path = self.report(run_path)
        workbook = load_workbook(file_path)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        email_phone_column = headers.index('email-phone') + 1
        password_column = headers.index('password') + 1

        for row in sheet.iter_rows(min_row=2, values_only=True):
            try:
                chrome_options = webdriver.ChromeOptions()
                chrome_options.add_argument("--disable-notifications")

                driver = webdriver.Chrome(options=chrome_options)
                driver.get(post)
                email_phone = row[email_phone_column - 1]
                password = row[password_column - 1]
                wait.pause(5)
                self.cookies(driver)
                self.email(email_phone, driver)
                self.password(password, driver)
                self.login_btn(driver)
                wait.pause(5)
                self.share(driver)
                self.share_now(driver, email_phone, report_path)
                wait.pause_log(pause_min, "Sharing again")
            except TimeoutException as e:
                print(f"Timeout Exception occurred: {e}")
                continue
            finally:
                if driver:
                    driver.quit()

    def report(self, path):
        today_date = datetime.datetime.now().strftime("%Y-%m-%d")
        run = 1
        file_name_template = f"run_{today_date}_{{}}.txt"
        while True:
            file_name = file_name_template.format(run)
            file_path = os.path.join(path, file_name)
            if not os.path.exists(file_path):
                with open(file_path, 'w'):
                    pass
                return file_path
            else:
                run += 1

    def cookies(self, driver):
        locator = (By.XPATH, "//span[text()='Allow all cookies']")
        util = Utils
        element = WebDriverWait(driver, 10).until(EC.presence_of_element_located(locator))
        util.find_clickable_parent(driver, element)
        print("cookies accepted")

    def email(self, email_phone, driver):
        locator = (By.XPATH, "//input[@dir='ltr' and @aria-invalid='false' and @type='text']")

        wait = WaitingUtils()
        wait.wait_until_visible(driver, locator)
        element = driver.find_element(*locator)
        element.send_keys(email_phone)
        print("email entered")

    def password(self, password, driver):
        locator = (By.XPATH, "//input[@dir='ltr' and @aria-invalid='false' and @type='password']")

        wait = WaitingUtils()
        wait.wait_until_visible(driver, locator)
        element = driver.find_element(*locator)
        element.send_keys(password)
        print("password entered")

    def login_btn(self, driver):
        locator = (By.XPATH, '//*[@id="login_popup_cta_form"]/div/div[5]/div/div')
        wait = WaitingUtils
        wait.wait_until_clickable(driver, locator)
        print("logged_in")

    def share(self, driver):
        utils = Utils
        locator = (By.XPATH, "//span[contains(text(), 'Share')]/..")

        element = WebDriverWait(driver, 10).until(EC.presence_of_element_located(locator))

        utils.scroll_to_element(driver, element)

        wait = WaitingUtils
        wait.wait_until_clickable(driver, locator)
        print("Share button pressed")

    def share_now(self, driver, user, report_path):
        locator = (By.XPATH, "//span[contains(text(), 'Share now (Friends)')]/..")
        wait = WaitingUtils
        wait.wait_until_clickable(driver, locator)
        print(f"shared by{user}")
        print(f"writing in: {report_path}")
        with open(report_path, "a") as file:
            file.write(f"shared by: {user}\n")
